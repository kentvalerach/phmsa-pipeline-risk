"""
PHMSA Annual Report XLSX → CSV Extractor
=========================================
Extracts individual Part sheets (J, K, M, H, A-D, etc.) from the full
annual_gas_transmission_gathering_YYYY.xlsx files into clean CSVs.

Usage:
    python extract_phmsa_parts.py

Configure INPUT_DIR and OUTPUT_DIR below, then run.
The script will scan for all xlsx files matching the naming pattern
and extract the specified sheets.

Requirements:
    pip install openpyxl

Author: Kent (ROSEN Project 4 — Survival Model)
Date: 2026-02-03
"""

import openpyxl
import csv
import os
import glob
import sys
from pathlib import Path

# ============================================================
# CONFIGURATION — EDIT THESE PATHS
# ============================================================

# Where your downloaded xlsx files are
INPUT_DIR = r"C:\Phmsa\annual_gt"

# Where to save the extracted CSVs
OUTPUT_DIR = r"C:\Phmsa\annual_gt\extracted_csvs"

# Which sheets to extract (sheet_name → output_suffix)
# Extracts Part J by default. Uncomment others as needed.
SHEETS_TO_EXTRACT = {
    "GT AR Part J":      "Part_J",
    "GT AR Part K":      "Part_K",
    "GT AR Part M":      "Part_M",
    "GT AR Part H":      "Part_H",
    "GT AR Part A to D": "Part_A_to_D",
    # "GT AR Part L":    "Part_L",       # Class location miles
    # "GT AR Part I":    "Part_I",       # Assessment methods
    # "GT AR Part F to G":"Part_F_to_G", # Commodity volumes
}

# xlsx filename pattern (glob)
FILE_PATTERN = "annual_gas_transmission_gathering_*.xlsx"

# Header row index (0-based). PHMSA files have 2 info rows, headers at row 2.
HEADER_ROW = 2

# ============================================================
# EXTRACTION LOGIC — NO NEED TO EDIT BELOW
# ============================================================

def extract_sheet(xlsx_path, sheet_name, output_path, header_row=2):
    """Extract one sheet from xlsx to CSV."""
    
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    
    if sheet_name not in wb.sheetnames:
        wb.close()
        return None, f"Sheet '{sheet_name}' not found"
    
    ws = wb[sheet_name]
    
    headers = None
    data_rows = 0
    
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = None
        
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < header_row:
                continue  # Skip info rows
            
            if i == header_row:
                # Header row
                headers = []
                for cell in row:
                    headers.append(str(cell) if cell is not None else '')
                writer = csv.writer(f)
                writer.writerow(headers)
                continue
            
            # Data row — convert None to empty string
            values = []
            for cell in row:
                if cell is None:
                    values.append('')
                else:
                    values.append(str(cell))
            
            writer.writerow(values)
            data_rows += 1
    
    wb.close()
    return data_rows, None


def find_xlsx_files(input_dir, pattern):
    """Find all matching xlsx files and extract year from filename."""
    
    files = {}
    search = os.path.join(input_dir, pattern)
    
    for path in sorted(glob.glob(search)):
        filename = os.path.basename(path)
        # Extract year: annual_gas_transmission_gathering_YYYY.xlsx
        parts = filename.replace('.xlsx', '').split('_')
        year = parts[-1] if parts[-1].isdigit() else None
        
        if year:
            files[int(year)] = path
    
    return files


def main():
    print("=" * 65)
    print("PHMSA Annual Report XLSX → CSV Extractor")
    print("=" * 65)
    
    # Create output directory
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    # Find files
    xlsx_files = find_xlsx_files(INPUT_DIR, FILE_PATTERN)
    
    if not xlsx_files:
        print(f"\n  ERROR: No xlsx files found matching:")
        print(f"    {os.path.join(INPUT_DIR, FILE_PATTERN)}")
        print(f"\n  Check INPUT_DIR and FILE_PATTERN in the script.")
        print(f"  Current INPUT_DIR: {INPUT_DIR}")
        sys.exit(1)
    
    print(f"\n  Input directory:  {INPUT_DIR}")
    print(f"  Output directory: {OUTPUT_DIR}")
    print(f"  Found {len(xlsx_files)} xlsx files: {sorted(xlsx_files.keys())}")
    print(f"  Sheets to extract: {list(SHEETS_TO_EXTRACT.values())}")
    
    # Extract
    print(f"\n{'Year':>6} {'Sheet':<20} {'Rows':>7} {'Status':<20} File")
    print(f"{'-'*6} {'-'*20} {'-'*7} {'-'*20} {'-'*30}")
    
    total_files = 0
    errors = []
    
    for year in sorted(xlsx_files.keys()):
        xlsx_path = xlsx_files[year]
        
        for sheet_name, suffix in SHEETS_TO_EXTRACT.items():
            output_name = f"GT_AR_{year}_{suffix}.csv"
            output_path = os.path.join(OUTPUT_DIR, output_name)
            
            try:
                rows, err = extract_sheet(
                    xlsx_path, sheet_name, output_path, 
                    header_row=HEADER_ROW
                )
                
                if err:
                    print(f"  {year}  {suffix:<20} {'—':>7} {'⚠️  ' + err:<20}")
                    errors.append(f"{year} {suffix}: {err}")
                else:
                    size_kb = os.path.getsize(output_path) / 1024
                    print(f"  {year}  {suffix:<20} {rows:>6,} {'✅':>3}  {size_kb:>6.0f} KB  {output_name}")
                    total_files += 1
                    
            except Exception as e:
                print(f"  {year}  {suffix:<20} {'—':>7} {'❌ ERROR':<20} {str(e)[:40]}")
                errors.append(f"{year} {suffix}: {str(e)}")
    
    # Summary
    print(f"\n{'=' * 65}")
    print(f"  Extracted: {total_files} CSV files to {OUTPUT_DIR}")
    if errors:
        print(f"  Warnings:  {len(errors)}")
        for e in errors:
            print(f"    - {e}")
    
    # Verify schema consistency
    print(f"\n  Verifying Part J schema consistency...")
    partj_files = sorted(glob.glob(os.path.join(OUTPUT_DIR, "GT_AR_*_Part_J.csv")))
    
    if partj_files:
        ref_headers = None
        ref_year = None
        all_match = True
        
        for pf in partj_files:
            with open(pf, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                headers = next(reader)
            
            year_str = os.path.basename(pf).split('_')[2]
            
            if ref_headers is None:
                ref_headers = headers
                ref_year = year_str
            elif headers != ref_headers:
                all_match = False
                print(f"    ⚠️  {year_str} differs from {ref_year}!")
                print(f"       {year_str}: {len(headers)} cols")
                print(f"       {ref_year}: {len(ref_headers)} cols")
        
        if all_match:
            print(f"    ✅ All {len(partj_files)} Part J files: identical schema ({len(ref_headers)} cols)")
    
    print(f"\n  Done! Upload the CSVs from {OUTPUT_DIR} to Claude.")
    print(f"  For the survival model, Part J is essential.")
    print(f"  Parts K, M, H add covariables for the full model.")


if __name__ == "__main__":
    main()